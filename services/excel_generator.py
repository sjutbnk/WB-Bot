import os
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any

class ExcelGenerator:
    """Сервис для генерации профессионально оформленных Excel отчетов."""

    @staticmethod
    def generate_supply_report(data: List[Dict[str, Any]], filepath: str) -> str:
        """Создает оформленный Excel-файл с расчетом поставок."""
        # Готовим список строк для датафрейма
        rows = []
        for item in data:
            dist = item.get("distribution", {})
            rows.append({
                "Артикул": item["sku"],
                "Товар (Бренд / Категория)": item["name"],
                "Остаток на складах": item["total_stock"],
                "Товары в пути": item["transit_stock"],
                "Продаж/день (14д)": round(item["speed_14"], 2),
                "Продаж/день (30д)": round(item["speed_30"], 2),
                "Оборачиваемость (дни)": "∞" if item["turnover_days"] == 999 else round(item["turnover_days"], 1),
                "Рекомендовано поставить": item["needed_total"],
                "Коледино": dist.get("Коледино", 0),
                "Казань": dist.get("Казань", 0),
                "Электросталь": dist.get("Электросталь", 0),
                "Краснодар": dist.get("Краснодар", 0),
                "Екатеринбург": dist.get("Екатеринбург", 0)
            })

        df = pd.DataFrame(rows)
        
        # Сохраняем в файл через ExcelWriter
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Рекомендации поставок")
            
            # Получаем объект листа для стилизации
            workbook = writer.book
            worksheet = writer.sheets["Рекомендации поставок"]
            
            # --- Стилизация ---
            # Цветовые заливки
            header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Темно-синий
            deficit_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid") # Светло-зеленый (нужна поставка)
            normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            # Шрифты
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            data_font = Font(name="Calibri", size=10)
            bold_data_font = Font(name="Calibri", size=10, bold=True)
            
            # Выравнивание
            center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left_align = Alignment(horizontal="left", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")
            
            # Границы
            thin_border_side = Side(border_style="thin", color="D9D9D9")
            border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
            
            # Стилизуем шапку таблицы
            worksheet.row_dimensions[1].height = 28
            for col_idx in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border
            
            # Заполняем стилями данные
            for row_idx in range(2, len(df) + 2):
                worksheet.row_dimensions[row_idx].height = 20
                needed_total = df.iloc[row_idx - 2]["Рекомендовано поставить"]
                
                # Если нужна поставка, подсветим всю строчку нежно-зеленым
                row_fill = deficit_fill if needed_total > 0 else normal_fill
                
                for col_idx in range(1, len(df.columns) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.fill = row_fill
                    cell.border = border
                    cell.font = bold_data_font if col_idx in [1, 8] else data_font
                    
                    # Выравнивание колонок
                    if col_idx in [1, 7]: # Артикул, оборачиваемость
                        cell.alignment = center_align
                    elif col_idx == 2: # Название товара
                        cell.alignment = left_align
                    else: # Цифровые значения
                        cell.alignment = right_align
                        # Форматирование чисел
                        if col_idx in [3, 4, 8, 9, 10, 11, 12, 13]:
                            cell.number_format = "#,##0"
                        elif col_idx in [5, 6]:
                            cell.number_format = "0.00"

            # Включаем сетку линий
            worksheet.views.sheetView[0].showGridLines = True
            
            # Автоматическая ширина столбцов с запасом
            for col in worksheet.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        # Учитываем перенос строки в заголовках
                        lines = str(cell.value).split('\n')
                        for line in lines:
                            max_len = max(max_len, len(line))
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        return filepath

    @staticmethod
    def generate_campaign_report(data: List[Dict[str, Any]], filepath: str) -> str:
        """Создает транспонированный Excel-отчет по рекламным кампаниям (показатели слева, артикулы сверху)."""
        # 1. Считаем итоги для первой колонки "Итого"
        total_budget = sum(item.get("Бюджет", 0) for item in data if item.get("Зоны показа") == "Весь период")
        total_views = sum(item.get("Показы", 0) for item in data if item.get("Зоны показа") == "Весь период")
        total_clicks = sum(item.get("Клики", 0) for item in data if item.get("Зоны показа") == "Весь период")
        total_carts = sum(item.get("Корзина", 0) for item in data if item.get("Зоны показа") == "Весь период")
        total_orders = sum(item.get("Заказы", 0) for item in data if item.get("Зоны показа") == "Весь период")
        total_orders_sum = sum(item.get("Заказы сумма", 0) for item in data if item.get("Зоны показа") == "Весь период")
        total_assoc_orders = sum(item.get("Ассоц. заказы, шт.", 0) for item in data if item.get("Зоны показа") == "Весь период")
        total_assoc_sum = sum(item.get("Ассоц. заказы, руб", 0) for item in data if item.get("Зоны показа") == "Весь период")
        total_profit = sum(item.get("Прогноз. приб. без опер. расх.", 0) for item in data if item.get("Зоны показа") == "Весь период")
        
        # Расчетные метрики для Итого
        total_ctr = (total_clicks / total_views * 100) if total_views else 0.0
        total_cpc = (total_budget / total_clicks) if total_clicks else 0.0
        total_cpm = (total_budget / total_views * 1000) if total_views else 0.0
        total_cr_to_cart = (total_carts / total_clicks * 100) if total_clicks else 0.0
        total_cr_to_order = (total_orders / total_carts * 100) if total_carts else 0.0
        total_cr = (total_orders / total_clicks * 100) if total_clicks else 0.0
        total_cpo = (total_budget / total_orders) if total_orders else None
        total_drr = (total_budget / total_orders_sum * 100) if total_orders_sum else None
        
        # Средние цены по кампаниям для Итого
        c_count = sum(1 for item in data if item.get("Зоны показа") == "Весь период" and item.get("Цена до СПП"))
        avg_price_before = sum(item.get("Цена до СПП", 0) for item in data if item.get("Зоны показа") == "Весь период" and item.get("Цена до СПП")) / c_count if c_count else 0
        avg_price_after = sum(item.get("Цена после СПП", 0) for item in data if item.get("Зоны показа") == "Весь период" and item.get("Цена после СПП")) / c_count if c_count else 0
        avg_spp_discount = (avg_price_before - avg_price_after) / avg_price_before * 100 if avg_price_before else 0
        avg_margin = (total_profit / total_orders_sum * 100) if total_orders_sum else (total_profit / avg_price_before * 100 if avg_price_before else 0)

        # Словарь Итого
        summary_row = {
            "Бюджет": round(total_budget, 2),
            "Показы": total_views,
            "Клики": total_clicks,
            "CTR": round(total_ctr, 2),
            "CPC": round(total_cpc, 2),
            "CPM": round(total_cpm, 2),
            "Корзина": total_carts,
            "Добавления в корзину": round(total_cr_to_cart, 2),
            "Заказы": total_orders,
            "Заказы сумма": round(total_orders_sum, 2),
            "Добавление в заказ": round(total_cr_to_order, 2),
            "Ассоц. заказы, шт.": total_assoc_orders,
            "Ассоц. заказы, руб": round(total_assoc_sum, 2),
            "CR": round(total_cr, 2),
            "CPO": round(total_cpo, 2) if total_cpo is not None else None,
            "ДРРз": round(total_drr, 2) if total_drr is not None else None,
            "Цена до СПП": round(avg_price_before, 2),
            "Цена после СПП": round(avg_price_after, 2),
            "Скидка МП": round(avg_spp_discount, 2),
            "Прогноз. марж": round(avg_margin, 2),
            "Прогноз. приб. без опер. расх.": round(total_profit, 2)
        }

        # Метрики по строкам (ключ, название, тип_форматирования)
        metrics_keys = [
            ("Бюджет", "Бюджет", "float"),
            ("Показы", "Показы", "int"),
            ("Клики", "Клики", "int"),
            ("CTR", "CTR", "float"),
            ("CPC", "CPC", "float"),
            ("CPM", "CPM", "float"),
            ("Корзина", "Корзина", "int"),
            ("Добавления в корзину", "Добавления в корзину", "float"),
            ("Заказы", "Заказы", "int"),
            ("Заказы сумма", "Заказы сумма", "float"),
            ("Добавление в заказ", "Добавление в заказ", "float"),
            ("Ассоц. заказы, шт.", "Ассоц. заказы, шт.", "int"),
            ("Ассоц. заказы, руб", "Ассоц. заказы, руб", "float"),
            ("CR", "CR", "float"),
            ("CPO", "CPO", "float"),
            ("ДРРз", "ДРРз", "float"),
            ("Цена до СПП", "Цена до СПП", "float"),
            ("Цена после СПП", "Цена после СПП", "float"),
            ("Скидка МП", "Скидка МП", "float"),
            ("Прогноз. марж", "Прогноз. марж", "float"),
            ("Прогноз. приб. без опер. расх.", "Прогноз. приб. без опер. расх.", "float")
        ]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Отчет по кампаниям"

        # Заливки и шрифты
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid") # Серый
        side_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Светло-серый
        normal_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        header_font = Font(name="Calibri", size=10, bold=True)
        bold_font = Font(name="Calibri", size=10, bold=True)
        data_font = Font(name="Calibri", size=10)
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        thin_border_side = Side(border_style="thin", color="D9D9D9")
        border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)

        # Пишем заголовки столбцов А и Б
        ws.cell(row=1, column=1, value="Показатель").alignment = left_align
        ws.cell(row=2, column=1, value="Зоны показа").alignment = left_align
        ws.cell(row=1, column=2, value="Итого").alignment = center_align
        ws.cell(row=2, column=2, value="Весь период").alignment = center_align

        ws.cell(row=1, column=1).fill = header_fill
        ws.cell(row=2, column=1).fill = header_fill
        ws.cell(row=1, column=2).fill = side_fill
        ws.cell(row=2, column=2).fill = side_fill

        ws.cell(row=1, column=1).font = header_font
        ws.cell(row=2, column=1).font = header_font
        ws.cell(row=1, column=2).font = header_font
        ws.cell(row=2, column=2).font = header_font

        # Заполняем показатели по строкам (Column A и Column B)
        for row_offset, (key, label, f_type) in enumerate(metrics_keys, 3):
            # Показатели (A)
            cell_a = ws.cell(row=row_offset, column=1, value=label)
            cell_a.alignment = left_align
            cell_a.fill = header_fill
            cell_a.font = bold_font
            cell_a.border = border
            
            # Итого (B)
            cell_b = ws.cell(row=row_offset, column=2, value=summary_row.get(key))
            cell_b.alignment = right_align
            cell_b.fill = side_fill
            cell_b.font = bold_font
            cell_b.border = border
            if cell_b.value is not None:
                cell_b.number_format = "#,##0" if f_type == "int" else "0.00"

        # Заполняем данные по кампаниям в столбцы C, D, E...
        col_idx = 3
        current_campaign = ""
        
        for item in data:
            name = item.get("Название")
            zone = item.get("Зоны показа")
            
            if name:
                current_campaign = name
                
            # Заголовки столбцов для РК
            c_cell_1 = ws.cell(row=1, column=col_idx, value=current_campaign)
            c_cell_1.alignment = center_align
            c_cell_1.font = header_font
            c_cell_1.fill = header_fill
            c_cell_1.border = border
            
            c_cell_2 = ws.cell(row=2, column=col_idx, value=zone)
            c_cell_2.alignment = center_align
            c_cell_2.font = header_font
            c_cell_2.fill = header_fill
            c_cell_2.border = border
            
            # Заполняем значения строк по текущему столбцу
            for row_offset, (key, label, f_type) in enumerate(metrics_keys, 3):
                val = item.get(key)
                cell = ws.cell(row=row_offset, column=col_idx, value=val)
                cell.alignment = right_align
                cell.font = data_font
                cell.border = border
                cell.fill = normal_fill
                
                if val is not None:
                    cell.number_format = "#,##0" if f_type == "int" else "0.00"
                    
            col_idx += 1

        # Объединяем ячейки с названием РК в Row 1 для каждого столбца РК (по 3 на каждую)
        num_campaigns = len(data) // 3
        for c_idx in range(num_campaigns):
            start_col = 3 + c_idx * 3
            end_col = start_col + 2
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            
            # Восстанавливаем границы после объединения
            for c in range(start_col, end_col + 1):
                ws.cell(row=1, column=c).border = border

        # Сетка линий
        ws.views.sheetView[0].showGridLines = True
        
        # Высоты строк
        ws.row_dimensions[1].height = 36
        ws.row_dimensions[2].height = 24
        for r in range(3, len(metrics_keys) + 3):
            ws.row_dimensions[r].height = 20

        # Автоматическая ширина колонок с ограничением по ширине названия
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value:
                    lines = str(cell.value).split('\n')
                    for line in lines:
                        max_len = max(max_len, len(line))
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, 11), 32)

        wb.save(filepath)
        return filepath
